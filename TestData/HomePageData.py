import openpyxl


class HomePageData:
    test_homepage_data = [{"name": "Amrita", "email": "Das@gmail.com", "pass": "Gaur1232", "gender": "Female"},
                          {"name": "Adhoksaja", "email": "aaa@mail.com", "pass": "pass123", "gender": "Male"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("D:\\Celebal Tech\\Selenium with Python\\Udemy Course\\PythonDemo.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            # to select testcase2 only
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["name"] = "Amrita"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
