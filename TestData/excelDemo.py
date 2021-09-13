import openpyxl

book = openpyxl.load_workbook("D:\\Celebal Tech\\Selenium with Python\\Udemy Course\\PythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=2, column=2)
print(cell.value)
sheet.cell(row=2, column=3).value = "Das@gmail.com"
print(sheet.cell(row=2, column=3).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A4'].value)

Dict = {}

for i in range(1, sheet.max_row + 1):
    # to select testcase2 only
    if sheet.cell(row=i, column=1).value == "testcase2":
        for j in range(2, sheet.max_column + 1):  # to get columns
        # Dict["name"] = "Amrita"
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
